#! /usr/bin/python3
import sys
from openpyxl import load_workbook, Workbook

#Coleta os dados da primeira planilha e os coloca em uma variavel#

wb = load_workbook(filename = "teste.xlsx")
pagina = wb['Foglio1']
contador = 2
dados = []
coluna = "A"
while contador <= 13:
	col = (pagina[coluna+str(contador)].value)
	dados.append(col)
	contador+=1
print (dados)

#Coleta os dados da segunda planilha e os coloca em uma variavel#

wb = load_workbook(filename = "teste2.xlsx")
pagina = wb['Foglio1']
dados2 = []
coluna = "A"
contador = 2
while contador <=13:
	col = (pagina[coluna+str(contador)].value)
	dados2.append(col)
	contador+=1
print (dados2)

#Cria um arquivo mestre e #



